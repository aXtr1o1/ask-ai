"""
Evaluation Runner
=================
Runs the Understanding Agent → Analysis Agent pipeline for each enabled question,
cycling through ALL questions once per round, for CYCLES rounds total.

  Total rows = len(enabled_questions) × CYCLES

Chat history is DISABLED by default — every question gets its own fresh
session UUID so each call is fully independent with no prior context.
Use --history to re-enable shared-session-per-cycle mode.

After all questions in one cycle complete, waits BREAK_SECONDS before the
next cycle.

Excel columns per row (one row = one run):
  Q#  | Cycle | Question
  ---- Understanding Agent ----
  UA Thought            (full streaming thought text)
  UA Module Selected    (list of FM modules)
  UA Summary            (query_summary)
  ---- Analysis Agent ----
  AA Thought            (full streaming thought text)
  AA Reasoning          (reasoning field from AnalysisOutput)
  AA Filter Fields      (filter_fields dict as JSON)
  AA Filter Values      (filter_values dict as JSON)

Usage:
  cd nanosoft-ai-backend
  python -m app.api.advance.evaluation.run_evaluation

Optional flags:
  --cycles  N   Override the number of cycles (default: 5)
  --break   N   Override the break seconds between cycles (default: 5)
  --history     Enable chat history (shared session per cycle). Default: OFF
"""
import argparse
import json
import logging
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from app.api.advance.Understanding_Agent.agent import classify_query
from app.api.advance.Understanding_Agent.conversation_memory import conversation_memory
from app.api.advance.analysis.agent import analyze_query
from app.api.advance.evaluation.questions import QUESTIONS

# =============================================================================
# Logging
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("evaluation.runner")


# =============================================================================
# Config
# =============================================================================
CYCLES        = 5    # number of full cycles through all questions
BREAK_SECONDS = 5    # pause between cycles
OUTPUT_DIR    = Path(__file__).parent  # same directory as this file


# =============================================================================
# Excel helpers
# =============================================================================
# Color palette
_HEADER_BG  = "1F2D3D"   # dark navy
_HEADER_FG  = "FFFFFF"

_UA_BG      = "1A3A5C"   # deep blue  (Understanding Agent header group)
_AA_BG      = "1A4A2E"   # deep green (Analysis Agent header group)

_ROW_ODD    = "F0F4F8"   # very light blue-grey
_ROW_EVEN   = "FFFFFF"   # white

_ERROR_BG   = "FDECEA"   # soft red for error rows
_ERROR_FG   = "B71C1C"


def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _make_cell(ws, row: int, col: int, value, *, bold=False, bg=None, fg="000000",
               wrap=False, h_align="left", v_align="top"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=fg, name="Calibri", size=10)
    cell.alignment = Alignment(wrap_text=wrap, horizontal=h_align,
                               vertical=v_align, shrink_to_fit=False)
    cell.border    = _thin_border()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def _build_headers(ws):
    """Write the two-row merged header and return column widths map."""

    # -- Row 1: group headers -------------------------------------------------
    group_headers = [
        (1, 3,  "General",             _HEADER_BG),
        (4, 6,  "Understanding Agent", _UA_BG),
        (7, 10, "Analysis Agent",      _AA_BG),
    ]
    for start_col, end_col, label, bg in group_headers:
        ws.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1,   end_column=end_col,
        )
        cell = ws.cell(row=1, column=start_col, value=label)
        cell.font      = Font(bold=True, color=_HEADER_FG, name="Calibri", size=11)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()

    # -- Row 2: individual column headers -------------------------------------
    col_headers = [
        # General
        ("Q #",           8),
        ("Cycle #",       8),
        ("Question",      55),
        # Understanding Agent
        ("UA - Thought\n(Full thinking text)",          55),
        ("UA - Module Selected",                        22),
        ("UA - Summary",                                45),
        # Analysis Agent
        ("AA - Thought\n(Full thinking text)",          55),
        ("AA - Reasoning",                              45),
        ("AA - Filter Fields\n(JSON)",                  40),
        ("AA - Filter Values\n(JSON)",                  40),
    ]

    col_bg_map = {
        1: _HEADER_BG, 2: _HEADER_BG, 3: _HEADER_BG,
        4: _UA_BG,     5: _UA_BG,     6: _UA_BG,
        7: _AA_BG,     8: _AA_BG,     9: _AA_BG,    10: _AA_BG,
    }

    for col_idx, (label, width) in enumerate(col_headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font      = Font(bold=True, color=_HEADER_FG, name="Calibri", size=10)
        cell.fill      = PatternFill("solid", fgColor=col_bg_map[col_idx])
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 38


def _write_row(ws, excel_row: int, q_id: int, cycle_num: int,
               question: str, ua: dict, aa: dict | None, error_msg: str | None):
    """Write one data row (one run result) into the worksheet."""

    is_error   = error_msg is not None
    bg_color   = _ERROR_BG if is_error else (_ROW_ODD if excel_row % 2 == 1 else _ROW_EVEN)
    text_color = _ERROR_FG if is_error else "000000"

    def cell(col, value, wrap=True):
        return _make_cell(ws, excel_row, col, value,
                          bg=bg_color, fg=text_color, wrap=wrap, v_align="top")

    cell(1,  q_id,      wrap=False)
    cell(2,  cycle_num, wrap=False)
    cell(3,  question)

    if is_error:
        cell(4,  f"ERROR: {error_msg}")
        for c in range(5, 11):
            cell(c, "")
    else:
        # Understanding Agent columns
        cell(4,  ua.get("thought", ""))
        cell(5,  ", ".join(ua.get("modules", [])) or "-")
        cell(6,  ua.get("query_summary", ""))

        # Analysis Agent columns
        cell(7,  aa.get("thought", "") if aa else "")
        cell(8,  aa.get("reasoning", "") if aa else "")
        cell(9,  json.dumps(aa.get("filter_fields", {}), indent=2) if aa else "")
        cell(10, json.dumps(aa.get("filter_values", {}), indent=2) if aa else "")

    ws.row_dimensions[excel_row].height = 80


# =============================================================================
# Core runner
# =============================================================================
def run_single(
    question: str,
    session_id: str,
    *,
    store_history: bool = False,
) -> tuple[dict, dict | None, str | None]:
    """
    Execute one UA -> AA pipeline call.

    Parameters
    ----------
    question      : the natural-language question to evaluate
    session_id    : UUID passed to classify_query so the UA can look up prior
                    context.  When chat history is OFF, callers pass a brand-new
                    UUID for every question so no context is ever found.
    store_history : when True, add_turn() is called after the run so the NEXT
                    call with the same session_id can see this result.
                    Set to False (default) for history-OFF mode.

    Returns (ua_result, aa_result, error_message).
    aa_result is None when UA intent is not db_query.
    error_message is None on success.
    """
    # -- Understanding Agent --------------------------------------------------
    try:
        ua = classify_query(
            query            = question,
            session_id       = session_id,
            thought_callback = None,          # batch mode
        )
    except Exception as exc:
        logger.error("UA failed: %s", exc)
        return {}, None, f"Understanding Agent error: {exc}"

    logger.info(
        "  [UA] intent=%s  modules=%s  summary=%.80s...",
        ua.get("intent"), ua.get("modules"), ua.get("query_summary", ""),
    )

    intent  = ua.get("intent", "general")
    summary = ua.get("query_summary", question)

    # -- Analysis Agent (only for db_query intent) ----------------------------
    aa = None
    if intent == "db_query" and ua.get("modules"):
        try:
            aa = analyze_query(
                query_summary    = summary,
                modules          = ua["modules"],
                thought_callback = None,   # batch mode
                last_db_turn     = None,   # AA does not need chat history directly
            )
        except Exception as exc:
            logger.error("AA failed: %s", exc)
            return ua, None, f"Analysis Agent error: {exc}"

        logger.info(
            "  [AA] reasoning=%.80s...  filter_fields=%s  filter_values=%s",
            aa.get("reasoning", ""),
            list(aa.get("filter_fields", {}).keys()),
            list(aa.get("filter_values", {}).keys()),
        )
    else:
        logger.info(
            "  [AA] skipped - intent=%s (not db_query or no modules)",
            intent,
        )

    # -- Optionally store turn into memory ------------------------------------
    # Only done when history is ON.  In history-OFF mode we skip this entirely
    # so that the next question truly starts from a blank slate.
    if store_history:
        conversation_memory.add_turn(
            session_id       = session_id,
            user_query       = question,
            query_summary    = summary,
            intent           = intent,
            modules          = ua.get("modules", []),
            filter_fields    = aa.get("filter_fields", {}) if aa else {},
            filter_values    = aa.get("filter_values", {}) if aa else {},
            general_response = ua.get("general_response") or "",
        )

    return ua, aa, None


def run_evaluation(
    cycles: int = CYCLES,
    break_secs: int = BREAK_SECONDS,
    *,
    history: bool = False,          # ← OFF by default
):
    enabled_questions = [q for q in QUESTIONS if q.get("enabled", True)]
    total_rows        = len(enabled_questions) * cycles
    history_label     = "ON (shared session per cycle)" if history else "OFF (fresh session per question)"

    logger.info("=" * 70)
    logger.info("Evaluation Runner started")
    logger.info("  Questions   : %d enabled / %d total", len(enabled_questions), len(QUESTIONS))
    logger.info("  Cycles      : %d  (all questions run once per cycle)", cycles)
    logger.info("  Total rows  : %d  (%d questions x %d cycles)",
                total_rows, len(enabled_questions), cycles)
    logger.info("  Chat history: %s", history_label)
    logger.info("  Break       : %ds between cycles", break_secs)
    logger.info("=" * 70)

    # -- Set up workbook ------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"
    ws.freeze_panes = "A3"   # freeze header rows

    _build_headers(ws)

    excel_row = 3   # data starts at row 3 (rows 1-2 are headers)
    row_count = 0

    # -- Main loop ------------------------------------------------------------
    #
    #  Outer loop  : questions
    #  Inner loop  : cycles (run the same question N times in a row)
    #
    # -------------------------------------------------------------------------
    for q_idx, q in enumerate(enabled_questions, start=1):
        q_id     = q["id"]
        question = q["question"]

        # When history is ON, generate ONE session id shared across all
        # cycles for this specific question.
        q_session_id = str(uuid.uuid4())

        logger.info("")
        logger.info("=" * 70)
        logger.info("Question %d/%d  [Q#%d] — history=%s",
                    q_idx, len(enabled_questions), q_id, "ON" if history else "OFF")
        logger.info("  %s", question)
        if history:
            logger.info("  Shared session: %s", q_session_id)
        logger.info("=" * 70)

        for cycle_num in range(1, cycles + 1):
            row_count += 1

            # History OFF → unique session per run (no context carried over)
            # History ON  → shared session across all runs of this question
            session_id = q_session_id if history else str(uuid.uuid4())

            logger.info("  --- Run %d/%d  (overall row %d/%d)  session=%s",
                        cycle_num, cycles, row_count, total_rows, session_id)

            ua, aa, err = run_single(
                question,
                session_id,
                store_history=history,
            )
            _write_row(ws, excel_row, q_id, cycle_num, question, ua, aa, err)
            excel_row += 1

        # Pause between different questions
        if q_idx < len(enabled_questions):
            logger.info("")
            logger.info("  Finished all runs for Q#%d. Pausing %ds before next question...",
                        q_id, break_secs)
            time.sleep(break_secs)

    # -- Save workbook --------------------------------------------------------
    history_tag = "history_on" if history else "history_off"
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"evaluation_results_{history_tag}_{timestamp}.xlsx"
    wb.save(output_path)

    logger.info("")
    logger.info("=" * 70)
    logger.info("Evaluation complete!")
    logger.info("   Rows written : %d", row_count)
    logger.info("   Saved to     : %s", output_path)
    logger.info("=" * 70)

    return str(output_path)


# =============================================================================
# Entry point
# =============================================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run UA -> AA pipeline evaluation")
    parser.add_argument(
        "--cycles",
        type=int,
        default=CYCLES,
        help=f"Number of full cycles through all questions (default: {CYCLES})",
    )
    parser.add_argument(
        "--break",
        dest="break_secs",
        type=int,
        default=BREAK_SECONDS,
        help=f"Break seconds between cycles (default: {BREAK_SECONDS})",
    )
    parser.add_argument(
        "--history",
        action="store_true",
        default=False,
        help="Enable chat history (shared session per cycle). Default: OFF (each question is isolated).",
    )
    args = parser.parse_args()

    run_evaluation(cycles=args.cycles, break_secs=args.break_secs, history=args.history)
